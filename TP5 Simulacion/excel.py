import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

def mostrar_excel(datos, coches,rk):
    def limpiar_array_internos(lista):
        nuevo_array = []
        for elemento in lista:
            if isinstance(elemento, list):
                for sub_elemento in elemento:
                    if isinstance(sub_elemento, list):
                        nuevo_array.extend(sub_elemento)
                    else:
                        nuevo_array.append(sub_elemento)
            else:
                nuevo_array.append(str(elemento) if isinstance(elemento, (int, float)) else elemento)
        return nuevo_array

    # Limpiando datos internos de los arrays
    nuevos_datos = []
    for subarray in datos:
        nuevos_datos.append(limpiar_array_internos(subarray))

    columnas = [
        "Evento",
        "Reloj",
        "RND Estadia",
        "Estadia",
        "Fin estacionamiento",
        "RND Proxima llegada",
        "Proxima llegada",
        "RND Valor C",
        "Valor C",
        "Tiempo de Cobro",
        "RND Tipo",
        "Tipo",
        "Estado playa",
        "Capacidad",
        "Porcentaje utilizacion",
        "Estado zona cobro",
        "Cola",
        "Cobro total",
        "Cobro acumulado"
    ]

    columnas_coches = [
        "Numero",
        "Tipo",
        "Estado",
        "Hora llegada",
        "Hora fin de estacionamiento",
        "Tiempo Cobro"
    ]

    wb = openpyxl.Workbook()
    hoja = wb.active

    # Crear encabezados combinados
    hoja.merge_cells("C1:E1")
    hoja["C1"] = "Auto"

    hoja.merge_cells("F1:G1")
    hoja["F1"] = "Proxima llegada"

    hoja.merge_cells("H1:J1")
    hoja["H1"] = "Cobro"

    hoja.merge_cells("K1:L1")
    hoja["K1"] = "Tipo"

    hoja.merge_cells("M1:O1")
    hoja["M1"] = "Playa"

    hoja.merge_cells("P1:S1")
    hoja["P1"] = "Cobro"

    # Añadir encabezados de coches
    column_coche = 20
    for inx in range(len(coches[-1])):
        columnas.extend(columnas_coches)
        letra_inicial = get_column_letter(column_coche)
        letra_final = get_column_letter(column_coche + 4)

        rango = f'{letra_inicial}1:{letra_final}1'
        hoja.merge_cells(rango)
        hoja[f'{letra_inicial}1'] = f"Coche Activo {inx}"
        column_coche += 6

    hoja.append(columnas)

    fila = 3
    for i, fila_datos in enumerate(nuevos_datos):
        for j, valor in enumerate(fila_datos):
            hoja.cell(row=fila, column=j+1).value = str(valor)
        
        if i < len(coches):
            coches_actuales = limpiar_array_internos(coches[i])
            for k, coche in enumerate(coches_actuales, start=20):
                hoja.cell(row=fila, column=k).value = str(coche)
        
        fila += 1

    # Dar formato a las celdas de encabezado
    for row in range(1, 3):
        for col in range(1, hoja.max_column + 1):
            celda = hoja.cell(row=row, column=col)
            celda.alignment = Alignment(horizontal='center')
            celda.font = Font(bold=True)

     # Crear la hoja para Runge-Kutta
    hoja_rk = wb.create_sheet(title="Runge-Kutta")

    encabezados_rk = ["t", "C", "k1", "C+k1/2", "k2", "C+k2/2", "k3", "C+k3", "k4", "Ci+1"]
    hoja_rk.append(encabezados_rk)

    # Añadir los datos de Runge-Kutta
    for fila_rk in rk:
        hoja_rk.append(fila_rk)

    # Dar formato a las celdas de encabezado de Runge-Kutta
    for col in range(1, len(encabezados_rk) + 1):
        celda = hoja_rk.cell(row=1, column=col)
        celda.alignment = Alignment(horizontal='center')
        celda.font = Font(bold=True)

    wb.save("datos.xlsx")



    wb.save("datos.xlsx")
