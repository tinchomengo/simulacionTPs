import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# Estructura de las filas
# 0 ["Evento"]
# 1 [Reloj (Minutos)]
# 2 [[RndEstadia,Estadia],[RndProxLlegada,ProxLlegada],[FinCobro]],
# 3 [RndAuto1,"tipo"],
# 4 [EstadoPlaya, CapacidadPlaya,PorcentajeUtilizacion],
# 5 [EstadoZonaCobro,ColaCobro],
# 6 [CobroTotal,CobroAcumulado]

def mostrar_excel(datos, coches):
    def limpiar_array_internos(lista):
        nuevo_array = []
        for elemento in lista:
            if isinstance(elemento, list):
                for sub_elemento in elemento:
                    if isinstance(sub_elemento, list):
                        nuevo_array.extend(sub_elemento)
                    else:
                        nuevo_array.append(sub_elemento)
            else:
                nuevo_array.append(str(elemento) if isinstance(elemento, (int, float)) else elemento)
        return nuevo_array

    nuevos_datos = []
    for subarray in datos:
        fila = []
        for elemento in subarray:
            if isinstance(elemento, list):
                for valor in elemento:
                    if isinstance(valor, list):
                        fila.extend(valor)
                    else:
                        fila.append(valor)
            else:
                fila.append(elemento)
        nuevos_datos.append(fila)


    nuevos_datos = limpiar_array_internos(nuevos_datos)



    columnas = [
        "Evento",
        "Reloj",
        "RND",
        "Estadia",
        "Fin estacionamiento",
        "RND",
        "Proxima llegada",
        "RND",
        "Tipo",
        "Estado playa",
        "Capacidad",
        "Porcentaje utilizacion",
        "Estado zona cobro",
        "Cola",
        "Cobro total",
        "Cobro acumulado"
    ]

    columnas_coches = [
        "Numero",
        "Tipo",
        "Estado",
        "Hora llegada",
        "Hora fin de estacionamiento"
    ]

    wb = openpyxl.Workbook()
    hoja = wb.active

    
    hoja.merge_cells("C1:I1")
    hoja["C1"] = "Auto"

    hoja.merge_cells("J1:K1")
    hoja["J1"] = "Playa"

    hoja.merge_cells("M1:P1")
    hoja["M1"] = "Cobro"

    column_coche = 17

    for inx in range(len(coches[-1])):
        columnas.extend(columnas_coches)
        letra_inicial = get_column_letter(column_coche)
        letra_final = get_column_letter(column_coche + 4)

        rango = f'{letra_inicial}1:{letra_final}1'
        hoja.merge_cells(rango)
        hoja[f'{letra_inicial}1'] = f"Coche Activo {inx}"
        column_coche += 5

    hoja.append(columnas)

    c = 1
    fila = 3
    i = 0

    for x in nuevos_datos:
        hoja.cell(row=fila, column=c).value = str(x)
        c += 1
        if c > 16:
            if i < len(coches):
                nuevo_coches = limpiar_array_internos(coches[i])
                for j, coche in enumerate(nuevo_coches, start=17):
                    hoja.cell(row=fila, column=j).value = str(coche)

            c = 1
            fila += 1
            i += 1

    for row in range(1,3):
        for col in range(1, hoja.max_column + 1):
            celda = hoja.cell(row=row, column=col)
            celda.alignment = Alignment(horizontal='center')
            celda.font = Font(bold=True)


    wb.save("datos.xlsx")
