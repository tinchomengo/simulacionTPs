import openpyxl

def mostrar_excel(datos,semanas,comision_promedio):

    def limpiar_array_internos(lista):
        nuevo_array = []
        for elemento in lista:
            if isinstance(elemento, list):
                for sub_elemento in elemento:
                    if isinstance(sub_elemento, list):
                        nuevo_array.extend(sub_elemento)
                    else:
                        nuevo_array.append(sub_elemento)
            else:
                nuevo_array.append(str(elemento) if isinstance(elemento, (int, float)) else elemento)
        return nuevo_array

    nuevos_datos = []
    for subarray in datos:
        fila = []
        for elemento in subarray:
            if isinstance(elemento, list):
                for valor in elemento:
                    if isinstance(valor, list):
                        fila.extend(valor)
                    else:
                        fila.append(valor)
            else:
                fila.append(elemento)
        nuevos_datos.append(fila)


    nuevos_datos=limpiar_array_internos(nuevos_datos)

    encabezados = [
        "Semana",

        "RndDemanda",
        "Demanda",
        "RndAuto1", 
        "TipoAuto1",
        "RndAuto2", 
        "TipoAuto2",
        "RndAuto3", 
        "TipoAuto3",
        "RndAuto4", 
        "TipoAuto4",
        "ComisionAuto1", 
        "CantidadComision1",
        "ComisionAuto2", 
        "CantidadComision2",
        "ComisionAuto3",
        "CantidadComision3",
        "ComisionAuto4", 
        "CantidadComision4",
        "RndSorteo", 
        "ResSorteo",
        "ComisionFila",
        "ComisionAcumulada",
        #        "ComisionPromedio",


        "RndDemanda",
        "Demanda",
        "RndAuto1", 
        "TipoAuto1",
        "RndAuto2", 
        "TipoAuto2",
        "RndAuto3", 
        "TipoAuto3",
        "RndAuto4", 
        "TipoAuto4",
        "ComisionAuto1", 
        "CantidadComision1",
        "ComisionAuto2", 
        "CantidadComision2",
        "ComisionAuto3",
        "CantidadComision3",
        "ComisionAuto4", 
        "CantidadComision4",
        "RndSorteo", 
        "ResSorteo",
        "ComisionFila",
        "ComisionAcumulada",
        #        "ComisionPromedio",


        "RndDemanda",
        "Demanda",
        "RndAuto1", 
        "TipoAuto1",
        "RndAuto2", 
        "TipoAuto2",
        "RndAuto3", 
        "TipoAuto3",
        "RndAuto4", 
        "TipoAuto4",
        "ComisionAuto1", 
        "CantidadComision1",
        "ComisionAuto2", 
        "CantidadComision2",
        "ComisionAuto3",
        "CantidadComision3",
        "ComisionAuto4", 
        "CantidadComision4",
        "RndSorteo", 
        "ResSorteo",
        "ComisionFila",
        "ComisionAcumulada",
#        "ComisionPromedio",


        "RndDemanda",
        "Demanda",
        "RndAuto1", 
        "TipoAuto1",
        "RndAuto2", 
        "TipoAuto2",
        "RndAuto3", 
        "TipoAuto3",
        "RndAuto4", 
        "TipoAuto4",
        "ComisionAuto1", 
        "CantidadComision1",
        "ComisionAuto2", 
        "CantidadComision2",
        "ComisionAuto3",
        "CantidadComision3",
        "ComisionAuto4", 
        "CantidadComision4",
        "RndSorteo", 
        "ResSorteo",
        "ComisionFila",
        "ComisionAcumulada",
#        "ComisionPromedio",


        "ComisionPromedio"
    ]

    wb = openpyxl.Workbook()
    hoja = wb.active
    hoja.append(encabezados)

    for i, dato in enumerate(semanas, start=2):
        hoja.cell(row=i, column=1).value = dato

    c = 2
    fila = 2
    for x in nuevos_datos:
        hoja.cell(row=fila, column=c).value = str(x)
        c += 1
        if c > 89:
           #93

            c = 2
            fila += 1

    for i, dato in enumerate(comision_promedio, start=2):
        hoja.cell(row=i, column=90).value = dato
#                          column=93
    wb.save("datos.xlsx")